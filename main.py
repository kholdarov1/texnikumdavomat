from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from typing import Optional
import os
import base64
import csv
import uuid
import math

app = FastAPI()

# 🔓 CORS – Netlify yoki boshqa domenlardan murojaat qilish uchun
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://davomat2026.netlify.app/"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


DATA_DIR = "data"
IMAGES_DIR = os.path.join(DATA_DIR, "images")
os.makedirs(IMAGES_DIR, exist_ok=True)

# 📍 ISHXONA KOORDINATASI
# Google Maps’dan aniq koordinatani olib, shu yerga qo'yasan:
# masalan: 39.764321, 64.432198
OFFICE_LAT = 40.68560701258604   # <-- bu yerga ishxonaning latitude (kenglik)
OFFICE_LNG = 71.90455733991762   # <-- bu yerga ishxonaning longitude (uzunlik)
MAX_DISTANCE_METERS = 50  # ishxona atrofidagi radius (metrda)


class CheckIn(BaseModel):
    image_base64: str          # frontenddan keladigan rasm (dataURL)
    mode: str                  # "arrival" yoki "departure"
    lat: Optional[float] = None
    lng: Optional[float] = None


def get_today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def get_today_csv_path() -> str:
    today = get_today_str()
    return os.path.join(DATA_DIR, f"davomat_{today}.csv")


def distance_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    """
    Yer yuzasidagi ikki nuqta orasidagi masofani metrda hisoblaydi (Haversine formulasi).
    """
    R = 6371000  # Yer radiusi (m)

    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)

    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(
        dlambda / 2
    ) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


# ✅ Davomat yozib borish (geolokatsiya bilan)
@app.post("/checkin")
def checkin(data: CheckIn):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # mode tekshirish
    mode = data.mode
    if mode not in ("arrival", "departure"):
        raise HTTPException(status_code=400, detail="Mode noto'g'ri")

    if mode == "arrival":
        is_late = now.hour > 8 or (now.hour == 8 and now.minute > 0)
        status = "KECHIKIB KELDI" if is_late else "O'z vaqtida keldi"
        mode_text = "Ishga keldi"
    else:
        is_late = False
        status = "Ishdan ketdi"
        mode_text = "Ishdan ketdi"

    # 🔐 Geolokatsiya tekshiruv
    if data.lat is None or data.lng is None:
        raise HTTPException(status_code=400, detail="Geolokatsiya talab qilinadi")

    dist = distance_m(data.lat, data.lng, OFFICE_LAT, OFFICE_LNG)
    if dist > MAX_DISTANCE_METERS:
        raise HTTPException(
            status_code=400,
            detail=f"Siz ish joyi hududida emassiz (taxminiy masofa {int(dist)} m).",
        )

    # 🖼 Rasmni faylga saqlash
    try:
        if "," in data.image_base64:
            _, b64data = data.image_base64.split(",", 1)
        else:
            b64data = data.image_base64

        img_bytes = base64.b64decode(b64data)
    except Exception:
        raise HTTPException(status_code=400, detail="Rasm formatida xato")

    filename = f"{date_str}_{time_str.replace(':', '-')}_{uuid.uuid4().hex[:6]}.jpg"
    img_path = os.path.join(IMAGES_DIR, filename)
    with open(img_path, "wb") as f:
        f.write(img_bytes)

    # 📄 CSV ga yozib borish
    os.makedirs(DATA_DIR, exist_ok=True)
    csv_path = get_today_csv_path()
    file_exists = os.path.isfile(csv_path)

    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(
                ["sana", "vaqt", "mode", "status", "image_file", "lat", "lng", "dist_m"]
            )
        writer.writerow(
            [
                date_str,
                time_str,
                mode_text,
                status,
                filename,
                data.lat,
                data.lng,
                round(dist, 2),
            ]
        )

    # Frontend uchun javob
    return {
        "date": date_str,
        "time": time_str,
        "mode_text": mode_text,
        "status": status,
        "is_late": is_late,
        "image_url": filename,
        "distance_m": int(dist),
    }


# 📤 Bugungi kun uchun Excel (.xlsx) tayyorlash va yuklab berish
@app.get("/export-today")
def export_today():
    today = get_today_str()
    csv_path = get_today_csv_path()

    if not os.path.exists(csv_path):
        raise HTTPException(status_code=404, detail="Bugungi davomat topilmadi")

    xlsx_path = os.path.join(DATA_DIR, f"davomat_{today}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"

    # Sarlavha qatori
    ws.append(["Rasm", "Sana", "Vaqt", "Amal", "Status", "Lat", "Lng", "Masofa (m)"])

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        row_idx = 2  # 1-qator sarlavha
        for row in reader:
            ws.cell(row=row_idx, column=2, value=row["sana"])
            ws.cell(row=row_idx, column=3, value=row["vaqt"])
            ws.cell(row=row_idx, column=4, value=row["mode"])
            ws.cell(row=row_idx, column=5, value=row["status"])

            # lat, lng, masofa ham yozamiz (agar CSV da bo'lsa)
            ws.cell(row=row_idx, column=6, value=row.get("lat"))
            ws.cell(row=row_idx, column=7, value=row.get("lng"))
            ws.cell(row=row_idx, column=8, value=row.get("dist_m"))

            img_file = os.path.join(IMAGES_DIR, row["image_file"])
            if os.path.exists(img_file):
                img = XLImage(img_file)
                img.width = 80
                img.height = 80
                ws.add_image(img, f"A{row_idx}")

            ws.row_dimensions[row_idx].height = 60
            row_idx += 1

    wb.save(xlsx_path)

    return FileResponse(
        path=xlsx_path,
        filename=os.path.basename(xlsx_path),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

